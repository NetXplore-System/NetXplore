import logging
import uuid
from typing import List
from pydantic import BaseModel

import networkx as nx
from community import community_louvain
from networkx.algorithms import community as nx_community

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import JSONResponse

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select

from database import get_db
from models import Research, ResearchFilter, NetworkAnalysis, Comparisons
from auth_router import get_current_user
from utils import apply_comparison_filters, find_common_nodes, mark_common_nodes, get_network_metrics


from fastapi.responses import StreamingResponse
import json
from datetime import datetime
from openpyxl import Workbook
from io import BytesIO


logger = logging.getLogger(__name__)
router = APIRouter()

@router.get("/history/analyze/compare")
async def analyze_network_comparison_history(
        research_id: str = Query(...),
        min_weight: int = Query(1),
        node_filter: str = Query(""),
        highlight_common: bool = Query(False),
        metrics: str = Query(None),
        comparison_index: int = Query(0),
        db: AsyncSession = Depends(get_db)
):
    try:
       
        research = await db.get(Research, research_id)
        if not research:
            raise HTTPException(status_code=404, detail="Research not found")
        
        
        analysis_query = select(NetworkAnalysis).where(NetworkAnalysis.research_id == research_id)
        original_result = await db.execute(analysis_query)
        original_data = original_result.scalars().first()

        comparison_query = select(Comparisons).where(Comparisons.research_id == research_id)
        comparison_result = await db.execute(comparison_query)
        comparison_data = comparison_result.scalars().all()

        if original_data:
            original_data = original_data.to_dict()
        if comparison_index < 0 or comparison_index >= len(comparison_data):
            raise HTTPException(status_code=404, detail="Comparison index out of range")

        specific_comparison = comparison_data[comparison_index].to_dict()

        filtered_original = apply_comparison_filters(original_data, node_filter, min_weight)
        filtered_comparison = apply_comparison_filters(specific_comparison, node_filter, min_weight)

        if highlight_common:
            common_nodes = find_common_nodes(filtered_original, filtered_comparison)
            mark_common_nodes(filtered_original, common_nodes)
            mark_common_nodes(filtered_comparison, common_nodes)

        return JSONResponse(content={
            "original": filtered_original,
            "comparison": filtered_comparison,
            "metrics": get_network_metrics(filtered_original, filtered_comparison, metrics)
        }, status_code=200)

    except Exception as e:
        logger.error(f"Error in network comparison: {e}")
        raise HTTPException(detail=str(e), status_code=500)



@router.get("/history/{user_id}")
async def get_user_history(
    user_id: str,
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    try:
        user_uuid = uuid.UUID(user_id)

        if str(user_uuid) != current_user["user_id"]:
            raise HTTPException(
                status_code=403,
                detail="Access forbidden: You can only view your own research history"
            )
        
        query = select(Research).where(Research.user_id == user_uuid)
        result = await db.execute(query)
        researches = result.scalars().all()
        
        history = []
        
        for research in researches:
            filter_query = select(ResearchFilter).where(
                ResearchFilter.research_id == research.research_id
            )
            filter_result = await db.execute(filter_query)
            filters = filter_result.scalars().first()
            
            analysis_query = select(NetworkAnalysis).where(
                NetworkAnalysis.research_id == research.research_id
            )
            analysis_result = await db.execute(analysis_query)
            analysis = analysis_result.scalars().first()
            
            comparison_query = select(Comparisons).where(
                Comparisons.research_id == research.research_id
            )
            comparison_result = await db.execute(comparison_query)
            comparisons = comparison_result.scalars().all()
            
           
            research_entry = {
                **research.to_dict(), 
                "filters": filters.to_dict() if filters else None,
                "analysis": analysis.to_dict() if analysis else None,
                "comparisons": [comp.to_dict() for comp in comparisons] if comparisons else []
            }
            
            history.append(research_entry) 
        
        return JSONResponse(
            content={
                "status": "success",
                "history": history,
            },
            status_code=200
        )
            
    except ValueError as ve:
        logger.error(f"Invalid UUID format: {ve}")
        raise HTTPException(
            status_code=400,
            detail="Invalid user ID format"
        )
    except Exception as e:
        logger.error(f"Error fetching user history: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Error fetching history: {str(e)}"
        )
        
class CommunityAnalysisData(BaseModel):
    nodes: List[dict]
    links: List[dict]
    algorithm: str = Query("louvain")

@router.post("/history/analyze/communities") 
async def analyze_communities_history(
        data: CommunityAnalysisData
):
    try:    
        algorithm = data.algorithm
    
        G = nx.Graph()

        for node in data.nodes:
            G.add_node(node["id"], **{k: v for k, v in node.items() if k != "id"})

        for link in data.links:
            source = link["source"]
            target = link["target"]
            weight = link.get("weight", 1)

            if isinstance(source, dict) and "id" in source:
                source = source["id"]
            if isinstance(target, dict) and "id" in target:
                target = target["id"]

            G.add_edge(source, target, weight=weight)

        communities = {}
        node_communities = {}

        if algorithm == "louvain":
            partition = community_louvain.best_partition(G)
            node_communities = partition

            for node, community_id in partition.items():
                if community_id not in communities:
                    communities[community_id] = []
                communities[community_id].append(node)

        elif algorithm == "girvan_newman":
            communities_iter = nx_community.girvan_newman(G)
            communities_list = list(next(communities_iter))

            for i, community in enumerate(communities_list):
                communities[i] = list(community)
                for node in community:
                    node_communities[node] = i

        elif algorithm == "greedy_modularity":
            communities_list = list(nx_community.greedy_modularity_communities(G))

            for i, community in enumerate(communities_list):
                communities[i] = list(community)
                for node in community:
                    node_communities[node] = i
        else:
            logger.error(f"Unknown algorithm: {algorithm}. Supported: louvain, girvan_newman, greedy_modularity")
            raise HTTPException(
                detail=f"Unknown algorithm: {algorithm}. Supported: louvain, girvan_newman, greedy_modularity",
                status_code=400
            )

        communities_list = [
            {
                "id": community_id,
                "size": len(nodes),
                "nodes": nodes,
                "avg_betweenness": sum(data.nodes[i]["betweenness"]
                                       for i, node in enumerate(data.nodes)
                                       if node["id"] in nodes) / len(nodes) if nodes else 0,
                "avg_pagerank": sum(data.nodes[i]["pagerank"]
                                    for i, node in enumerate(data.nodes)
                                    if node["id"] in nodes) / len(nodes) if nodes else 0,
            }
            for community_id, nodes in communities.items()
        ]

        communities_list.sort(key=lambda x: x["size"], reverse=True)

        for i, node in enumerate(data.nodes):
            node_id = node["id"]
            if node_id in node_communities:
                data.nodes[i]["community"] = node_communities[node_id]

        return JSONResponse(content={
            "nodes": data.nodes,
            "links": data.links,
            "communities": communities_list,
            "node_communities": node_communities,
            "algorithm": algorithm,
            "num_communities": len(communities),
            "modularity": community_louvain.modularity(node_communities, G) if algorithm == "louvain" else None
        }, status_code=200)

    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        logger.error(f"Error in community detection: {e}")
        raise HTTPException(detail=str(e), status_code=500)
 

def format_date_safe(date_val):
    if not date_val:
        return ""
    if isinstance(date_val, datetime):
        return date_val.strftime("%Y-%m-%d")
    try:
        return datetime.fromisoformat(date_val).strftime("%Y-%m-%d")
    except Exception:
        return str(date_val)



@router.get("/export/excel/{research_id}")
async def export_excel_file(
    research_id: str,
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    research = await db.get(Research, research_id)
    if not research or str(research.user_id) != current_user["user_id"]:
        raise HTTPException(status_code=403, detail="Unauthorized")

    filters_result = await db.execute(select(ResearchFilter).where(ResearchFilter.research_id == research_id))
    filters = filters_result.scalars().first()

    analysis_result = await db.execute(select(NetworkAnalysis).where(NetworkAnalysis.research_id == research_id))
    analysis = analysis_result.scalars().first()

    comparisons_result = await db.execute(select(Comparisons).where(Comparisons.research_id == research_id))
    comparisons = comparisons_result.scalars().all()

    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append([
        "Research Name", "Platform", "Created At",
        "Start Date", "End Date", "Message Count",
        "Nodes", "Links", "Metric", "Communities"
    ])
    ws_summary.append([
        research.research_name,
        research.platform,
        research.created_at.strftime("%Y-%m-%d"),
        filters.start_date if filters else "",
        filters.end_date if filters else "",
        filters.min_messages if filters else "",
        len(analysis.nodes) if analysis and analysis.nodes else "",
        len(analysis.links) if analysis and analysis.links else "",
        analysis.metric_name if analysis else "",
        len(analysis.communities) if analysis and analysis.communities else ""
    ])

    ws_summary.append([])
    ws_summary.append([
        "Original File", "Comparison File",
        "Original Nodes", "Comparison Nodes", "Node Diff", "Node %",
        "Original Edges", "Comparison Edges", "Edge Diff", "Edge %",
        "Common Nodes", "Original Density", "Comparison Density"
    ])

    for comp in comparisons:
        stats = json.loads(comp.statistics or "{}")
        ws_summary.append([
            getattr(comp, "original_file_name", ""),
            getattr(comp, "file_name", ""),
            stats.get("original_node_count", ""),
            stats.get("comparison_node_count", ""),
            stats.get("node_difference", ""),
            stats.get("node_change_percent", ""),
            stats.get("original_link_count", ""),
            stats.get("comparison_link_count", ""),
            stats.get("link_difference", ""),
            stats.get("link_change_percent", ""),
            stats.get("common_nodes_count", ""),
            stats.get("original_density", ""),
            stats.get("comparison_density", "")
        ])

    ws_nodes = wb.create_sheet(title="Nodes")
    node_headers = ["id", "name", "group", "degree", "messages", "pagerank", "closeness", "betweenness", "eigenvector"]
    ws_nodes.append(node_headers)

    for node in analysis.nodes:
        ws_nodes.append([node.get(h, "") for h in node_headers])

    ws_links = wb.create_sheet(title="Links")
    ws_links.append(["source", "target", "weight"])
    for link in analysis.links:
        ws_links.append([
            link.get("source", ""),
            link.get("target", ""),
            link.get("weight", "")
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=research_{research_id}.xlsx"
        }
    )
